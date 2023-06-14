#!/usr/bin/env python3

import argparse
import logging
import os
import sys

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main():
    """Main function for CSV to Excel transformation and formatting tool.

    Takes CSV files from a specified directory, converts them to Excel format, applies colorization and formatting,
    and saves the resulting data as an Excel file. Logging level and certain actions are controlled by environment
    variables.

    Args:
        argv (list): Command-line arguments.
    """
    # Get logging level, set manually when running pipeline
    loglevel = os.environ.get("LOGLEVEL", "INFO").upper()
    if loglevel == "DEBUG":
        logging.basicConfig(
            level=loglevel,
            format="%(levelname)s [%(filename)s:%(lineno)d]: %(message)s",
        )
        logging.debug("Log level set to debug")
    else:
        logging.basicConfig(level=loglevel, format="%(levelname)s: %(message)s")
        logging.info("Log level set to info")

    # Process command-line arguments
    csv_dir = ""
    output_file = ""

    parser = argparse.ArgumentParser(
        description="Gather csv directory path and name of justification spreadsheet"
    )
    parser.add_argument("-i", "--csv-dir", dest="csv", help="Directory for scan csvs")
    parser.add_argument(
        "-o", "--output-file", dest="output", help="Name for justification excel file"
    )
    args = parser.parse_args()

    # csv_dir is the directory of the scan csvs, output_file is final xlsx file with justifications and coloring
    csv_dir = args.csv
    output_file = args.output

    # Convert all csvs to excel sheets
    # Generates two .xlsx spreadsheets, one with justifications (output_file) and one without justifications (all_scans.xlsx)
    convert_to_excel(csv_dir, output_file)
    workbook = openpyxl.load_workbook(output_file)
    # Colorize justifications for output_file
    _colorize_full(workbook)
    _set_all_column_widths(workbook)
    if os.environ["CI_COMMIT_BRANCH"] not in ["development", "master"]:
        _add_sheet_banners(workbook)
    workbook.save(output_file)


def _add_sheet_banners(workbook):
    """Apply colorization to all relevant worksheets in the workbook.

    Args:
        workbook (openpyxl.workbook): The Excel workbook object to apply colorization to.
    """
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        worksheet.insert_rows(1)
        cell = worksheet.cell(row=1, column=1)
        cell.value = "THESE FINDINGS ARE UNOFFICIAL AS THEY HAVE BEEN GENERATED ON A BRANCH OTHER THAN DEVELOPMENT OR MASTER. ONLY JUSTIFICATIONS FOR DEVELOPMENT OR MASTER BRANCH FINDINGS ARE CONSIDERED OFFICIAL."
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(
            start_color="00ff00ff", end_color="00ff00ff", fill_type="solid"
        )
        worksheet.merge_cells("A1:Z1")


def _colorize_full(workbook):
    """Returns the column index of a given value in the Excel worksheet.

    Args:
        sheet (openpyxl.worksheet): The Excel worksheet to search in.
        value (str): The column name to find.

    Returns:
        int: The column index.

    Raises:
        SystemExit: If the specified value is not found in the worksheet.
    """
    _colorize_sheet(workbook["Anchore CVE Results"])
    _colorize_sheet(workbook["Anchore Compliance Results"])
    _colorize_sheet(workbook["Twistlock Vulnerability Results"])

    if not os.environ.get("SKIP_OPENSCAP"):
        _colorize_sheet(workbook["OpenSCAP - DISA Compliance"])


def _get_column_index(sheet, value):
    """Apply colorization to a single worksheet based on cell values.

    Args:
        sheet (openpyxl.worksheet): The Excel worksheet to apply colorization to.
    """
    justification_column = None
    for i, col in enumerate(sheet.columns):
        if col[0].value == value:
            justification_column = i + 1
            break

    if not justification_column:
        logging.error("Could not find '%s' column", value)
        sys.exit(1)

    return justification_column


def _colorize_sheet(sheet):
    """Colorize justifications column."""
    justification_column = _get_column_index(sheet=sheet, value="Justification")
    results_column = None
    if sheet.title == "OpenSCAP - DISA Compliance":
        results_column = _get_column_index(sheet=sheet, value="result")
    for row in range(1, sheet.max_row + 1):
        justification_cell = sheet.cell(row=row, column=justification_column)
        # Apply appropriate highlighting to justification cell
        result = sheet.cell(row=row, column=results_column) if results_column else None
        if (not result or result.value == "fail") and justification_cell.value is None:
            # Fill cell in yellow
            justification_cell.fill = PatternFill(
                start_color="00ffff00", end_color="00ffff00", fill_type="solid"
            )
        elif justification_cell.value == "Inherited from base image.":
            # Fill cell in green
            justification_cell.fill = PatternFill(
                start_color="0000b050", end_color="0000b050", fill_type="solid"
            )
        elif justification_cell.value == "See Anchore CVE Results sheet":
            # Fill cell in gray
            justification_cell.fill = PatternFill(
                start_color="96969696", end_color="96969696", fill_type="solid"
            )
        elif justification_cell.value and justification_cell.value != "Justification":
            # Fill cell in blue
            justification_cell.fill = PatternFill(
                start_color="0000b0f0", end_color="0000b0f0", fill_type="solid"
            )


# convert all csvs to Excel file
# Generates output_file (w/ justifications) and all_scans.xlsx (w/o justifications)
def convert_to_excel(csv_dir, justification_sheet):
    """Convert CSV files from a given directory to an Excel file.

    Args:
        csv_dir (str): The path to the directory containing the CSV files.
        justification_sheet (str): The name of the Excel file to save the data to.
    """
    read_sum = pd.read_csv(csv_dir + "summary.csv")
    read_oscap = pd.read_csv(csv_dir + "oscap.csv")
    read_tl = pd.read_csv(csv_dir + "tl.csv")
    read_security = pd.read_csv(csv_dir + "anchore_security.csv")
    read_gates = pd.read_csv(csv_dir + "anchore_gates.csv")
    # column slice, remove last column which is justification to keep all_scans.xlsx with normal format
    read_oscap_no_justifications = read_oscap.iloc[:, :-1]
    read_tl_no_justifications = read_tl.iloc[:, :-1]
    read_security_no_justifications = read_security.iloc[:, :-1]
    read_gates_no_justifications = read_gates.iloc[:, :-1]
    # create all_scan.xlsx file (no justification or coloring used)
    with pd.ExcelWriter(csv_dir + "all_scans.xlsx") as writer:  # pylint: disable=E0110
        read_sum.to_excel(writer, sheet_name="Summary", header=True, index=False)
        read_oscap_no_justifications.to_excel(
            writer, sheet_name="OpenSCAP - DISA Compliance", header=True, index=False
        )
        read_tl_no_justifications.to_excel(
            writer,
            sheet_name="Twistlock Vulnerability Results",
            header=True,
            index=False,
        )
        read_security_no_justifications.to_excel(
            writer, sheet_name="Anchore CVE Results", header=True, index=False
        )
        read_gates_no_justifications.to_excel(
            writer, sheet_name="Anchore Compliance Results", header=True, index=False
        )
    with pd.ExcelWriter(justification_sheet) as writer:  # pylint: disable=E0110
        read_sum.to_excel(writer, sheet_name="Summary", header=True, index=False)
        read_oscap.to_excel(
            writer, sheet_name="OpenSCAP - DISA Compliance", header=True, index=False
        )
        read_tl.to_excel(
            writer,
            sheet_name="Twistlock Vulnerability Results",
            header=True,
            index=False,
        )
        read_security.to_excel(
            writer, sheet_name="Anchore CVE Results", header=True, index=False
        )
        read_gates.to_excel(
            writer, sheet_name="Anchore Compliance Results", header=True, index=False
        )


def _set_column_width(sheet, column_value, width, wrap=False):
    """Set column width and enable text wrap."""
    column = _get_column_index(sheet=sheet, value=column_value)
    sheet.column_dimensions[get_column_letter(column)].width = width
    if wrap:
        for cell in sheet[get_column_letter(column)]:
            cell.alignment = Alignment(wrap_text=True)


def _set_all_column_widths(workbook):
    """Set the width of specified columns in all relevant worksheets in the
    workbook.

    Args:
        workbook (openpyxl.workbook): The Excel workbook object to modify.
    """
    if not os.environ.get("SKIP_OPENSCAP"):
        openscap_disa = workbook["OpenSCAP - DISA Compliance"]
        _set_column_width(
            openscap_disa, column_value="scanned_date", width=20
        )  # scanned_date
        _set_column_width(
            openscap_disa, column_value="Justification", width=30
        )  # justification

    twistlock = workbook["Twistlock Vulnerability Results"]
    _set_column_width(twistlock, column_value="id", width=25)  # CVE
    _set_column_width(twistlock, column_value="packageName", width=20)  # packageName
    _set_column_width(
        twistlock, column_value="packageVersion", width=20
    )  # packageVersion
    _set_column_width(twistlock, column_value="vecStr", width=45)  # vecStr
    _set_column_width(
        twistlock, column_value="Justification", width=100
    )  # justification

    anchore_cve = workbook["Anchore CVE Results"]
    _set_column_width(anchore_cve, column_value="cve", width=25)  # CVE
    _set_column_width(anchore_cve, column_value="url", width=60)  # url
    _set_column_width(
        anchore_cve, column_value="Justification", width=100
    )  # justification

    anchore_compliance = workbook["Anchore Compliance Results"]
    _set_column_width(
        anchore_compliance, column_value="whitelist_name", width=30
    )  # whitelist_name
    _set_column_width(
        anchore_compliance, column_value="check_output", width=75
    )  # check_output
    _set_column_width(
        anchore_compliance, column_value="Justification", width=100
    )  # justification


if __name__ == "__main__":
    main(sys.argv[1:])  # pylint disable=E1121
