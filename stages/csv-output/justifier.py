#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

import yaml
import gitlab
import pathlib
import git
import os
import shutil
import json
import fnmatch
import sys
import getopt
import logging

##### The InheritableTriggerIds variable contains a list of Anchore compliance trigger_ids
##### that are inheritable by child images.
inheritableTriggerIds = [
    "639f6f1177735759703e928c14714a59",
    "c2e44319ae5b3b040044d8ae116d1c2f",
    "698044205a9c4a6d48b7937e66a6bf4f",
    "463a9a24225c26f7a5bf3f38908e5cb3",
    "bcd159901fe47efddae5c095b4b0d7fd",
    "320a97c6816565eedf3545833df99dd0",
    "953dfbea1b1e9d5829fbed2e390bd3af",
    "e7573262736ef52353cde3bae2617782",
    "addbb93c22e9b0988b8b40392a4538cb",
    "3456a263793066e9b5063ada6e47917d",
    "3e5fad1c039f3ecfd1dcdc94d2f1f9a0",
    "abb121e9621abdd452f65844954cf1c1",
    "34de21e516c0ca50a96e5386f163f8bf",
    "c4ad80832b361f81df2a31e5b6b09864",
]


def _load_local_hardening_manifest():
    """
    Load up the hardening_manifest.yaml file as a dictionary. Search for the file in
    the immediate repo first, if that is not found then search for the generated file.

    If neither are found then return None and let the calling function handle the error.

    """
    artifacts_path = os.environ["ARTIFACT_STORAGE"]
    paths = [
        pathlib.Path("hardening_manifest.yaml"),
        # Check for the generated hardening manifest. This method will be deprecated.
        pathlib.Path(artifacts_path, "preflight", "hardening_manifest.yaml"),
    ]

    for path in paths:
        logging.debug(f"Looking for {path}")
        if path.is_file():
            logging.debug(f"Using {path}")
            with path.open("r") as f:
                return yaml.safe_load(f)
        else:
            logging.debug(f"Couldn't find {path}")
    return None


def _load_remote_hardening_manifest(project, branch="master"):
    """
    Load up a hardening_manifest.yaml from a remote repository.

    If the manifest file is not found then None is returned. A warning will print
    to console to communicate which repository does not have a hardening manifest.

    """
    if project == "":
        return None

    logging.debug(f"Attempting to load hardening_manifest from {project}")

    try:
        gl = gitlab.Gitlab(os.environ["REPO1_URL"])
        proj = gl.projects.get(f"dsop/{project}", lazy=True)
        logging.debug(f"Connecting to dsop/{project}")

        hardening_manifest = proj.files.get(
            file_path="hardening_manifest.yaml", ref=branch
        )
        return yaml.safe_load(hardening_manifest.decode())

    except gitlab.exceptions.GitlabError:
        logging.info(
            "Could not load hardening_manifest. Defaulting backwards compatibility."
        )
        logging.warning(
            f"This method will be deprecated soon, please switch {project} to hardening_manifest.yaml"
        )

    except yaml.YAMLError as e:
        logging.error("Could not load the hardening_manifest.yaml")
        logging.error(e)
        sys.exit(1)

    return None


def _next_ancestor(image_path, greylist, hardening_manifest=None):
    """
    Grabs the parent image path from the current context. Will initially attempt to load
    a new hardening manifest and then pull the parent image from there. Otherwise it will
    default to the old method of using the greylist.

    If neither the hardening_manifest.yaml or the greylist field can be found then there
    is a weird mismatch during migration that needs further inspection.

    """

    # Try to get the parent image out of the local hardening_manifest.
    if hardening_manifest:
        return hardening_manifest["args"]["BASE_IMAGE"]

    # Try to load the hardening manifest from a remote repo.
    hm = _load_remote_hardening_manifest(project=image_path)
    if hm is not None:
        logging.debug(hm["args"]["BASE_IMAGE"])
        return hm["args"]["BASE_IMAGE"]

    try:
        logging.debug("using greylist for image parent name")
        return greylist["image_parent_name"]
    except KeyError as e:
        logging.error("Looks like a hardening_manifest.yaml cannot be found")
        logging.error(
            "Looks like the greylist has been updated to remove fields that should be present in hardening_manifest.yaml"
        )
        logging.error(e)
        sys.exit(1)


##### Clone the dccscr-whitelist repository
def cloneWhitelist(whitelistDir, whitelistRepo):
    # Delete the dccscr-whitelist folder (if it exists)
    if os.path.exists(whitelistDir):
        try:
            shutil.rmtree(whitelistDir)
        except OSError as e:
            print("Error: %s : %s" % (whitelistDir, e.strerror))

    # Clone the dccscr-whitelists repo
    dccscrWhitelistBranch = os.getenv("WL_TARGET_BRANCH")
    # Clone the dccscr-whitelists repo
    git.Repo.clone_from(
        whitelistRepo,
        os.path.join("./", "dccscr-whitelists"),
        branch=dccscrWhitelistBranch,
    )


# ##### Get the greylist for the source image
# def getSourceImageGreylistFile(whitelistDir, sourceImage):
#     sourceImageGreylistFile = ""
#     files = os.listdir(whitelistDir + "/" + sourceImage)
#     for file in files:
#         if fnmatch.fnmatch(file, "*.greylist"):
#             sourceImageGreylistFile = whitelistDir + "/" + sourceImage + "/" + file
#             print(sourceImageGreylistFile)
#     return sourceImageGreylistFile


def _get_greylist_file_contents(image_path, branch):
    """
    Grab the contents of a greylist file. Takes in the path to the image and
    determines the appropriate greylist.

    """
    greylist_file_path = f"{image_path}/{image_path.split('/')[-1]}.greylist"
    try:
        gl = gitlab.Gitlab(os.environ["REPO1_URL"])
        proj = gl.projects.get("dsop/dccscr-whitelists", lazy=True)
        f = proj.files.get(file_path=greylist_file_path, ref=branch)

    except gitlab.exceptions.GitlabError:
        logging.error(
            f"Whitelist retrieval attempted: {greylist_file_path} on {branch}"
        )
        logging.error(f"Error retrieving whitelist file: {sys.exc_info()[1]}")
        sys.exit(1)

    try:
        contents = json.loads(f.decode())
    except ValueError as e:
        logging.error("Could not load greylist as json")
        logging.error(e)
        sys.exit(1)

    if (
        contents["approval_status"] != "approved"
        and os.environ.get("CI_COMMIT_BRANCH").lower() == "master"
    ):
        logging.error("Unapproved image running on master branch")
        sys.exit(1)

    return contents


def _connect_to_db():
    """
    @return mariadb connection
    """
    conn = None
    try:
        conn = mysql.connector.connect(
            host=os.environ["vat_db_host"],
            database=os.environ["vat_db_database_name"],
            user=os.environ["vat_db_connection_user"],
            passwd=os.environ["vat_db_connection_pass"],
        )
        if conn.is_connected():
            # there are many connections to db so this should be uncommented
            # for troubleshooting
            logging.debug(
                "Connected to the host %s with user %s",
                os.environ["vat_db_host"],
                os.environ["vat_db_connection_user"],
            )
        else:
            logging.critical("Failed to connect to DB")
            sys.exit(1)
    except Error as err:
        logging.critical(err)
        if conn is not None and conn.is_connected():
            conn.close()
        sys.exit(1)

    return conn


def _vat_vuln_query(im_name, im_version):
    conn = None
    result = None
    try:
        conn = _connect_to_db()
        cursor = conn.cursor(buffered=True)
        # TODO: add new scan logic
        query = """SELECT c.name as container
                , c.version
                , CASE WHEN cl.type is NULL THEN "Pending" ELSE cl.type END as container_approval_status
                , f.finding
                , f.scan_source
                , f.in_current_scan
                , CASE
                WHEN fl.type is NULL THEN "Pending"
                WHEN cl.date_time is NULL and fl.type = "Approve" THEN "Reviewed"
                WHEN fl.date_time > cl.date_time and fl.type = "Approve" THEN "Reviewed"
                ELSE fl.type END as finding_status
                ,fl.text as approval_comments
                , fl2.text as justification
                , sr.description
                , f.package
                , f.package_path
                FROM findings_approvals f
                INNER JOIN containers c on f.imageid = c.id
                LEFT JOIN findings_log fl on f.id = fl.approval_id
                AND fl.id in (SELECT max(id) from findings_log WHERE type != "Justification" group by approval_id)
                LEFT JOIN findings_log fl2 on f.id = fl2.approval_id
                AND fl2.id in (SELECT max(id) from findings_log WHERE type = "Justification" group by approval_id)
                LEFT JOIN container_log cl on c.id = cl.imageid
                AND cl.id in (SELECT max(id) from container_log group by imageid)
                LEFT JOIN scan_results sr on c.id = sr.imageid AND f.finding = sr.finding AND f.package = sr.package AND f.package_path = sr.package_path
                AND jenkins_run in (SELECT max(jenkins_run) from scan_results WHERE imageid = c.id AND finding = f.finding AND package = f.package)
                WHERE f.inherited_id is NULL
                AND c.name=%s and c.version=%s
                AND f.in_current_scan = 1;"""
        cursor.execute(query, (im_name, im_version))
        result = cursor.fetchall()
    except Error as error:
        logging.info(error)
    finally:
        if conn is not None and conn.is_connected():
            conn.close()
    return result


def _get_vulns_from_query(row):
    vuln_dict = {}
    vuln_dict["whitelist_source"] = row[0]
    vuln_dict["version"] = row[1]
    vuln_dict["vulnerability"] = row[3]
    vuln_dict["vuln_source"] = row[4]
    vuln_dict["status"] = row[6]
    # vuln_dict["vuln_description"] =
    # vuln_dict["justification"] =
    # logging.debug(vuln_dict)
    return vuln_dict


def _get_complete_whitelist_for_image(image_name, whitelist_branch, hardening_manifest):
    """
    Pull all whitelisted CVEs for an image. Walk through the ancestry of a given
    image and grab all of vulnerabilities in the greylist associated with w layer.

    """
    total_whitelist = list()

    # TODO: remove after 30 day hardening_manifest merge cutoff
    greylist = _get_greylist_file_contents(
        image_path=image_name, branch=whitelist_branch
    )
    logging.info(f"Grabbing CVEs for: {image_name}")
    # get cves from vat
    result = _vat_vuln_query(os.environ["IMAGE_NAME"], os.environ["IMAGE_VERSION"])
    # parse CVEs from VAT query
    # empty list is returned if no entry or no cves. NoneType only returned if error.
    logging.info(result[0])
    if result is None:
        logging.error("No results from vat. Fatal error.")
        sys.exit(1)
    else:
        for row in result:
            vuln_dict = _get_vulns_from_query(row)
            if vuln_dict["status"] and vuln_dict["status"].lower() == "approve":
                total_whitelist.append(Vuln(vuln_dict, image_name))
                logging.debug(vuln_dict)
            else:
                logging.debug("There is no approval status present in result.")

    logging.debug(
        "Length of total whitelist for source image: " + str(len(total_whitelist))
    )

    #
    # Use the local hardening manifest to get the first parent. From here *only* the
    # the master branch should be used for the ancestry.
    #
    parent_image = _next_ancestor(
        image_path=image_name, greylist=greylist, hardening_manifest=hardening_manifest
    )

    # get parent cves from VAT
    while parent_image:
        logging.info(f"Grabbing CVEs for: {parent_image}")
        # TODO: remove this after 30 day hardening_manifest merge cutoff
        greylist = _get_greylist_file_contents(
            image_path=parent_image, branch=whitelist_branch
        )

        # TODO: swap this for hardening manifest after 30 day merge cutoff
        result = _vat_vuln_query(greylist["image_name"], greylist["image_tag"])
        logging.debug(result[0])
        # logging.debug(result[0])
        for row in result:
            vuln_dict = _get_vulns_from_query(row)
            if vuln_dict["status"] and vuln_dict["status"].lower() == "approve":
                total_whitelist.append(vuln_dict)

        parent_image = _next_ancestor(
            image_path=parent_image,
            greylist=greylist,
        )

    logging.info(f"Found {len(total_whitelist)} total whitelisted CVEs")
    return total_whitelist


# ##### Get the greylist file for all the base images
# ## TODO: Remove this
# def getAllGreylistFiles(
#     whitelistDir, sourceImage, sourceImageGreylistFile, hardening_manifest
# ):
#     # extract base_image from sourceImage .greylist file
#     baseImage = ""

#     # Add source image greylist file to allFiles
#     allFiles.append(sourceImageGreylistFile)

#     # Load the greylist file to pass to _next_ancestor
#     with open(sourceImageGreylistFile) as f:
#         try:
#             data = json.load(f)
#         except ValueError as e:
#             print("Error processing file: " + sourceImageGreylistFile, file=sys.stderr)
#             sys.exit(1)

#     # Check for empty .greylist file
#     if os.stat(sourceImageGreylistFile).st_size == 0:
#         print("Source image greylist file is empty")

#     # Get first parent image from hardening_manifest
#     else:
#         baseImage = _next_ancestor(
#             image_path=sourceImage,
#             greylist=data,
#             hardening_manifest=hardening_manifest,
#         )
#         if len(baseImage) == 0:
#             print("No parent image")

#     print("The following base image greylist files have been identified...")

#     # Add all parent image greylists to allFiles
#     # Break loop when it finds the last parent image
#     while len(baseImage) != 0:
#         files = os.listdir(whitelistDir + "/" + baseImage)
#         for file in files:
#             if fnmatch.fnmatch(file, "*.greylist"):
#                 baseImageGreylistFile = whitelistDir + "/" + baseImage + "/" + file
#                 print(baseImageGreylistFile)
#                 allFiles.append(baseImageGreylistFile)
#                 with open(baseImageGreylistFile) as f:
#                     try:
#                         data = json.load(f)
#                     except ValueError as e:
#                         print("Error processing file: " + file, file=sys.stderr)
#                 if os.stat(baseImageGreylistFile).st_size == 0:
#                     print("Source image greylist file is empty")
#                     baseImage = ""
#                 # return base image, checking hardening manifest first, then greylist. If no BASE_IMAGE, exit
#                 else:
#                     baseImage = _next_ancestor(image_path=baseImage, greylist=data)
#                     logging.debug(baseImage)

#     return allFiles


##### Read all greylist files and process into dictionary object
def getJustifications(total_whitelist, sourceImageName):

    cveOpenscap = {}
    cveTwistlock = {}
    cveAnchore = {}

    # Loop through all the greylist files
    # Getting results from VAT, just loop all findings, check if finding is related to base_images or source image
    for file in allFiles:
        # Load file into JSON object, print an error if the file doesn't load
        with open(file) as f:
            try:
                data = json.load(f)
            except ValueError as e:
                print("Error processing file: " + file, file=sys.stderr)

        # Check to see if the application is in approved status
        if "approval_status" in data.keys() and data["approval_status"] == "approved":
            # Get a list of all the whitelisted findings
            findings = data["whitelisted_vulnerabilities"]

            # Loop through the findings and create the corresponding dict object based on the vuln_source
            for finding in findings:
                if finding["status"] == "approved":

                    if "vulnerability" in finding.keys():
                        openscapID = finding["vulnerability"]
                        cveID = (
                            finding["vulnerability"] + "-" + finding["vuln_description"]
                        )
                        trigger_id_inherited = finding["vulnerability"]
                        trigger_id = finding["vulnerability"]

                        # Twistlock finding
                        if finding["vuln_source"] == "Twistlock":
                            if file == sourceImageGreylistFile:
                                cveTwistlock[cveID] = finding["justification"]
                            else:
                                cveTwistlock[cveID] = "Inherited from base image."

                        # Anchore finding
                        elif finding["vuln_source"] == "Anchore":
                            if file == sourceImageGreylistFile:
                                cveAnchore[cveID] = finding["justification"]
                                cveAnchore[trigger_id] = finding["justification"]
                            else:
                                cveAnchore[cveID] = "Inherited from base image."
                                if trigger_id in inheritableTriggerIds:
                                    cveAnchore[
                                        trigger_id
                                    ] = "Inherited from base image."

                        # OpenSCAP finding
                        elif finding["vuln_source"] == "OpenSCAP":
                            if file == sourceImageGreylistFile:
                                cveOpenscap[openscapID] = finding["justification"]
                            else:
                                cveOpenscap[openscapID] = "Inherited from base image."
            # print(cveAnchore)
    return cveOpenscap, cveTwistlock, cveAnchore


##### Process Openscap compliance justifications
def justificationsOpenscap(wb, justifications):
    # Process OpenSCAP - DISA Compliance tab
    sheet = wb["OpenSCAP - DISA Compliance"]
    for r in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=r, column=5)
        if cell.value == "identifiers":
            cell = sheet.cell(row=r, column=10)
            cell.value = "Justification"
        else:
            id = cell.value
            cell_justification = sheet.cell(row=r, column=10)

            if id in justifications.keys():
                cell_justification.value = justifications[id]

            # Apply appropriate highlighting to justification cell
            if cell_justification.value == None:
                cell_justification.fill = PatternFill(fill_type=None)
            elif cell_justification.value == "Inherited from base image.":
                cell_justification.fill = PatternFill(
                    start_color="0000b050", end_color="0000b050", fill_type="solid"
                )
            else:
                cell_justification.fill = PatternFill(
                    start_color="0000b0f0", end_color="0000b0f0", fill_type="solid"
                )


##### Process Twistlock justifications
def justificationsTwistlock(wb, justifications):
    # Process Twistlock Vulnerability Results tab
    sheet = wb["Twistlock Vulnerability Results"]
    for r in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=r, column=1)
        if cell.value == "id":
            cell = sheet.cell(row=r, column=10)
            cell.value = "Justification"
        else:
            cell2 = sheet.cell(row=r, column=3)
            cell_justification = sheet.cell(row=r, column=10)
            if cell2.value == None:
                id = cell.value
            else:
                id = cell.value + "-" + cell2.value

            if id in justifications.keys():
                cell_justification.value = justifications[id]

            cell3 = sheet.cell(row=r, column=5)
            cell4 = sheet.cell(row=r, column=6)
            id = cell.value + "-" + cell3.value + "-" + cell4.value

            if id in justifications.keys():
                cell_justification.value = justifications[id]

            # Apply appropriate highlighting to justification cell
            if cell_justification.value == None:
                cell_justification.fill = PatternFill(
                    start_color="00ffff00", end_color="00ffff00", fill_type="solid"
                )
            elif cell_justification.value == "Inherited from base image.":
                cell_justification.fill = PatternFill(
                    start_color="0000b050", end_color="0000b050", fill_type="solid"
                )
            else:
                cell_justification.fill = PatternFill(
                    start_color="0000b0f0", end_color="0000b0f0", fill_type="solid"
                )


##### Process Anchore justifications
def justificationsAnchore(wb, justifications):
    sheet = wb["Anchore CVE Results"]
    for r in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=r, column=2)
        if cell.value == "cve":
            cell = sheet.cell(row=r, column=9)
            cell.value = "Justification"
        else:
            cell2 = sheet.cell(row=r, column=4)
            cell_justification = sheet.cell(row=r, column=9)
            id = cell.value + "-" + cell2.value

            if id in justifications.keys():
                cell_justification.value = justifications[id]

            # Apply appropriate highlighting to justification cell
            if cell_justification.value == None:
                cell_justification.fill = PatternFill(
                    start_color="00ffff00", end_color="00ffff00", fill_type="solid"
                )
            elif cell_justification.value == "Inherited from base image.":
                cell_justification.fill = PatternFill(
                    start_color="0000b050", end_color="0000b050", fill_type="solid"
                )
            else:
                cell_justification.fill = PatternFill(
                    start_color="0000b0f0", end_color="0000b0f0", fill_type="solid"
                )


##### Process Anchore compliance justifications
def justificationsAnchoreComp(wb, justifications, inheritableTriggerIds):
    sheet = wb["Anchore Compliance Results"]
    for r in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=r, column=3)
        if cell.value == "trigger_id":
            cell = sheet.cell(row=r, column=13)
            cell.value = "Justification"
        else:
            cell2 = sheet.cell(row=r, column=5)
            cell_justification = sheet.cell(row=r, column=13)
            if cell2.value == "package":
                cell_justification.value = "See Anchore CVE Results sheet"

            id = cell.value
            if id in justifications.keys():
                cell_justification.value = justifications[id]

            # Apply appropriate highlighting to justification cell
            if cell_justification.value == None:
                cell_justification.fill = PatternFill(
                    start_color="00ffff00", end_color="00ffff00", fill_type="solid"
                )
            elif cell_justification.value == "Inherited from base image.":
                cell_justification.fill = PatternFill(
                    start_color="0000b050", end_color="0000b050", fill_type="solid"
                )
            elif cell_justification.value == "See Anchore CVE Results sheet":
                cell_justification.fill = PatternFill(
                    start_color="96969696", end_color="96969696", fill_type="solid"
                )
            else:
                cell_justification.fill = PatternFill(
                    start_color="0000b0f0", end_color="0000b0f0", fill_type="solid"
                )


def setColumnWidth(sheet, column, width, wrap=False):
    """Set column width and enable text wrap"""
    sheet.column_dimensions[get_column_letter(column)].width = width
    if wrap:
        for cell in sheet[get_column_letter(column)]:
            cell.alignment = Alignment(wrap_text=True)


def setAllColumnWidths(wb):

    openscap_disa = wb["OpenSCAP - DISA Compliance"]
    setColumnWidth(openscap_disa, column=9, width=20)  # scanned_date
    setColumnWidth(openscap_disa, column=10, width=30)  # justification

    twistlock = wb["Twistlock Vulnerability Results"]
    setColumnWidth(twistlock, column=1, width=25)  # CVE
    setColumnWidth(twistlock, column=5, width=20)  # packageName
    setColumnWidth(twistlock, column=6, width=20)  # packageVersion
    setColumnWidth(twistlock, column=9, width=45)  # vecStr
    setColumnWidth(twistlock, column=10, width=100)  # justification

    anchore_cve = wb["Anchore CVE Results"]
    setColumnWidth(anchore_cve, column=2, width=25, wrap=False)  # CVE
    setColumnWidth(anchore_cve, column=7, width=60)  # url
    setColumnWidth(anchore_cve, column=9, width=100)  # justification

    anchore_compliance = wb["Anchore Compliance Results"]
    setColumnWidth(
        anchore_compliance, column=12, width=30, wrap=False
    )  # whitelist_name
    setColumnWidth(
        anchore_compliance, column=13, width=100, wrap=False
    )  # justification
    setColumnWidth(anchore_compliance, column=6, width=75)  # check_output


##### Main function
def main(argv, inheritableTriggerIds):
    # Get logging level, set manually when running pipeline
    loglevel = os.environ.get("LOGLEVEL", "INFO").upper()
    if loglevel == "DEBUG":
        logging.basicConfig(
            level=loglevel,
            format="%(levelname)s [%(filename)s:%(lineno)d]: %(message)s",
        )
        logging.debug("Log level set to debug")
    else:
        logging.basicConfig(level=loglevel, format="%(levelname)s: %(message)s")
        logging.info("Log level set to info")
    # Process command-line arguments
    sourceFile = ""
    outputFile = ""
    sourceImage = ""
    try:
        opts, args = getopt.getopt(
            argv, "hi:o:s:", ["sourcefile=", "outputfile=", "sourceImage="]
        )
    except getopt.GetoptError:
        print("justifier.py -i <sourceFile> -o <outputfile> -s <sourceImage>")
        sys.exit(2)

    for opt, arg in opts:
        if opt == "-h":
            print("justifier.py -i <sourceFile> -o <outputfile> -s <sourceImage>")
            sys.exit()
        elif opt in ("-i", "--sourcefile"):
            sourceFile = arg
        elif opt in ("-o", "--outputfile"):
            outputFile = arg
        elif opt in ("-s", "--sourceImage"):
            sourceImage = arg

    print("Source file is " + sourceFile)
    print("Output file is " + outputFile)
    print("Source image is " + sourceImage)

    # sourceImageGreylistFile = ""
    # global allFiles
    # allFiles = []

    # TODO: Remove these commented out lines
    # Whitelist directory
    # global whitelistDir
    # whitelistDir = "dccscr-whitelists"

    # # Clone the whitelist repository
    # print("Cloning the dccscr-whitelists repository... ", end="", flush=True)
    # cloneWhitelist(whitelistDir, "https://repo1.dsop.io/dsop/dccscr-whitelists.git")
    # print("done.")

    # print("Getting source image greylist... ", end="", flush=True)
    # sourceImageGreylistFile = getSourceImageGreylistFile(whitelistDir, sourceImage)
    # print("done.")

    hardening_manifest = _load_local_hardening_manifest()
    if hardening_manifest is None:
        logging.error("Please update your project to contain a hardening_manifest.yaml")

    # TODO: Change this message
    print(
        "Getting greylist files for all parent images of " + sourceImage + "\n",
        end="",
        flush=True,
    )
    image_name = hardening_manifest["name"]
    wl_branch = os.getenv("WL_TARGET_BRANCH", default="master")

    total_whitelist = _get_complete_whitelist_for_image(
        image_name, wl_branch, hardening_manifest
    )

    # may need logic for hardening_manifest not being recovered if hardening_manifest == none etc.
    # Query vat for all whitelisted vulnerabilities
    # allFiles = getAllGreylistFiles(
    #     whitelistDir, sourceImage, sourceImageGreylistFile, hardening_manifest
    # )
    print("done.")

    # Get all justifications
    print("Gathering list of all justifications... ", end="", flush=True)
    jOpenscap, jTwistlock, jAnchore = getJustifications(
        total_whitelist, os.environ["IMAGE_NAME"]
    )
    print("done.")

    # Open the Excel file of the application we are updating
    wb = openpyxl.load_workbook(sourceFile)

    # Apply OpenSCAP compliance justifications
    print("Processing OpenSCAP Compliance Results... ", end="", flush=True)
    try:
        justificationsOpenscap(wb, jOpenscap)
    except:
        print("Unable to provide OpenSCAP Compliance justifications.")
    print("done.")

    # Apply Twistlock justifications
    print("Processing Twistlock Vulnerability Results... ", end="", flush=True)
    try:
        justificationsTwistlock(wb, jTwistlock)
    except:
        print("Unable to provide TL justifications.")
    print("done.")

    # Apply Anchore justifications
    print("Processing Anchore CVE Results... ", end="", flush=True)
    try:
        justificationsAnchore(wb, jAnchore)
    except:
        print("Unable to apply Anchore CVE justifications.")
    print("done.")

    # Apply Anchore compliance justifications
    print("Processing Anchore Compliance Results... ", end="", flush=True)
    try:
        justificationsAnchoreComp(wb, jAnchore, inheritableTriggerIds)
    except:
        print("Unable to apply Anchore Compliance justifications.")
    print("done.")

    print("Formatting... ", end="", flush=True)
    setAllColumnWidths(wb)
    print("done.")

    # Save the Excel file
    wb.save(outputFile)


if __name__ == "__main__":
    main(sys.argv[1:], inheritableTriggerIds)
